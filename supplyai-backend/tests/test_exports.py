"""Excel 导出 API 测试 — POST /exports/*.

设计:
- /sku-list 触发任务,返回 task_id
- 小数据量(< 5000 行)同步生成,status=success 时直接含 download_url
- /status 轮询任务状态
- /download 通过 task_id 流式下载文件
"""
from __future__ import annotations

import io

import openpyxl
from httpx import AsyncClient


async def test_export_sku_list_creates_task(client: AsyncClient) -> None:
    """触发导出返回 task_id 与状态."""
    resp = await client.post(
        "/api/supplyai/exports/sku-list",
        json={"tenant_id": 100228},
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["task_id"].startswith("EXP-")
    # 48 行 < 5000,同步完成
    assert data["status"] in {"success", "running", "pending"}
    assert data["row_count"] == 48


async def test_export_status_returns_state(client: AsyncClient) -> None:
    create = await client.post(
        "/api/supplyai/exports/sku-list",
        json={"tenant_id": 100228},
    )
    task_id = create.json()["task_id"]

    resp = await client.post(
        "/api/supplyai/exports/status",
        json={"tenant_id": 100228, "task_id": task_id},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["task_id"] == task_id
    assert data["status"] in {"pending", "running", "success", "failed"}


async def test_export_download_returns_xlsx(client: AsyncClient) -> None:
    """导出完成后下载 xlsx,内容能被 openpyxl 读取."""
    create = await client.post(
        "/api/supplyai/exports/sku-list",
        json={"tenant_id": 100228},
    )
    task_id = create.json()["task_id"]
    assert create.json()["status"] == "success"  # 同步完成

    resp = await client.post(
        "/api/supplyai/exports/download",
        json={"tenant_id": 100228, "task_id": task_id},
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument"
    )
    # 校验是合法 xlsx
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    # 表头 + 48 行数据
    assert ws.max_row == 49
    headers = [c.value for c in ws[1]]
    for must in ("MSKU", "ASIN", "店铺", "风险等级", "建议数量"):
        assert must in headers, f"表头缺 {must}"


async def test_export_download_filters_by_priority(client: AsyncClient) -> None:
    create = await client.post(
        "/api/supplyai/exports/sku-list",
        json={"tenant_id": 100228, "priorities": ["p1"]},
    )
    assert create.json()["row_count"] == 12  # seed 12 P1


async def test_export_status_404_unknown_task(client: AsyncClient) -> None:
    resp = await client.post(
        "/api/supplyai/exports/status",
        json={"tenant_id": 100228, "task_id": "EXP-NONEXISTENT"},
    )
    assert resp.status_code == 404
    assert resp.json()["detail"]["code"] == "EXPORT_TASK_NOT_FOUND"


async def test_export_download_404_unknown_task(client: AsyncClient) -> None:
    resp = await client.post(
        "/api/supplyai/exports/download",
        json={"tenant_id": 100228, "task_id": "EXP-NONEXISTENT"},
    )
    assert resp.status_code == 404
