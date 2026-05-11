"""API 旅程测试 — 业务需求拆成可执行的多端点串联场景.

⚠️ 这不是真浏览器 E2E。所有请求都通过 httpx + ASGITransport 进程内直发,
   不加载 React,不点 DOM。真浏览器 E2E 在 tests/test_browser_e2e.py。

每个测试模拟一段用户操作的"业务请求路径",串多个端点,验证:
  - 链路连通(每一步 200 OK)
  - 跨端点数据一致(同 calc_run_id / 状态机 / 公式)
  - 权限/边界(模型不能越权 tenant、未确认不落库)

约定:
  - 一个 test = 一个 User Story,docstring 用 As a / I want / So that 格式
  - 不重复测单端点细节(那是 test_skus.py 等的事),只测跨端点贯通
  - 用真实 seed 数据 + 注入的 AI stub(conftest 已 autouse)
"""
from __future__ import annotations

import io
import math
from dataclasses import dataclass, field
from typing import Any

import openpyxl
import pytest
import pytest_asyncio
from httpx import AsyncClient
from sqlalchemy import delete

from supplyai.db import async_session_factory
from supplyai.domain.ai.client import ChatResponse, ToolCall
from supplyai.models.mk import MkListingProductSources

TENANT_ID = 100228
SEED_CALC_RUN_ID = "DEMO-20260509-080000"


# ════════════════════════════════════════════════════════
# Epic 1 — 工作台日常监控
# 角色:运营负责人 / 主管
# 痛点:每天上岗第一件事,要 5 秒内看清今天工作量
# ════════════════════════════════════════════════════════


class TestEpic1WorkbenchMonitoring:
    async def test_us_1_1_operator_sees_risk_distribution_at_login(
        self, client: AsyncClient
    ) -> None:
        """US-1.1: 一打开工作台,看到全局风险分布.

        As a  运营负责人
        I want 登录后立刻看到 P1/P2/P3/safe 各多少 SKU
        So that 我能判断今天要处理多少紧急库存
        """
        snap = await client.post(
            "/api/supplyai/dashboard/snapshot", json={"tenant_id": TENANT_ID}
        )
        assert snap.status_code == 200
        data = snap.json()
        counts = data["risk_counts"]
        # 必须有 4 个等级
        assert set(counts) == {"p1", "p2", "p3", "safe"}
        # 加起来 = 该批次 FBA SKU 总数(seed 48)
        assert sum(counts.values()) == 48

    async def test_us_1_2_dashboard_shows_seven_day_stockout_count(
        self, client: AsyncClient
    ) -> None:
        """US-1.2: 一眼看出"几个 SKU 7 天内会断货".

        As a  运营负责人
        I want 工作台 hero 区直接显示"7 天内断货 X 个"
        So that 我不用先去过滤列表才知道紧急程度
        """
        snap = await client.post(
            "/api/supplyai/dashboard/snapshot", json={"tenant_id": TENANT_ID}
        )
        data = snap.json()
        assert "stockout_7_count" in data
        # 与 list 筛选 stockout_within_days=7 数量一致(口径自洽)
        listed = await client.post(
            "/api/supplyai/skus/list",
            json={
                "tenant_id": TENANT_ID,
                "stockout_within_days": 7,
                "page_size": 100,
            },
        )
        assert data["stockout_7_count"] == listed.json()["total"]

    async def test_us_1_3_risk_queue_top_p1_first(self, client: AsyncClient) -> None:
        """US-1.3: 高风险队列 P1 优先,带 action_hint.

        As a  运营负责人
        I want 队列前几行就是当前最紧急、能直接采取动作的 SKU
        So that 我能逐行处理而不用自己排优先级
        """
        q = await client.post(
            "/api/supplyai/dashboard/risk-queue",
            json={"tenant_id": TENANT_ID, "limit": 8},
        )
        rows = q.json()["rows"]
        # 至少前几行都应是 P1
        assert rows[0]["priority"] == "p1"
        assert rows[0]["action_hint"] == "urgent_purchase"
        # 排序单调:priority 不能从弱回到强
        order = {"p1": 0, "p2": 1, "p3": 2, "safe": 3}
        seen = -1
        for r in rows:
            cur = order[r["priority"]]
            assert cur >= seen
            seen = cur


# ════════════════════════════════════════════════════════
# Epic 2 — 从风险洞察到决策
# 角色:运营负责人
# 流程:列表筛 P1 → 进详情 → 看曲线/AI 解释 → 决定是否采购
# ════════════════════════════════════════════════════════


class TestEpic2RiskInvestigation:
    async def test_us_2_1_filter_p1_then_drill_into_detail(
        self, client: AsyncClient
    ) -> None:
        """US-2.1: 从备货列表筛 P1 → 进任一行详情.

        As a  运营负责人
        I want 在列表上勾"P1 紧急"标签,点任一行进详情页
        So that 我能针对性地决定这个 SKU 怎么补
        """
        listed = await client.post(
            "/api/supplyai/skus/list",
            json={"tenant_id": TENANT_ID, "priorities": ["p1"], "page_size": 50},
        )
        rows = listed.json()["rows"]
        assert all(r["priority"] == "p1" for r in rows)
        assert len(rows) == 12  # seed 12 个 P1

        # 选第一个进详情
        listing_id = rows[0]["id"]
        detail = await client.post(
            "/api/supplyai/skus/detail",
            json={"tenant_id": TENANT_ID, "listing_id": listing_id},
        )
        assert detail.status_code == 200
        # 详情必须含 summary + 预测序列
        d = detail.json()
        assert d["summary"]["priority"] == "p1"
        assert len(d["forecast_trend"]) >= 30

    async def test_us_2_2_detail_provides_history_and_forecast_for_chart(
        self, client: AsyncClient
    ) -> None:
        """US-2.2: 详情页拿到历史销量 + 未来预测,渲染曲线.

        As a  运营负责人
        I want 详情页能展示过去 90 天销量与未来 45 天预测的曲线
        So that 我能直观判断销量趋势,验证系统给的可售天数合理
        """
        first = await client.post(
            "/api/supplyai/skus/list",
            json={"tenant_id": TENANT_ID, "page_size": 1},
        )
        listing_id = first.json()["rows"][0]["id"]

        trends = await client.post(
            "/api/supplyai/skus/trends",
            json={"tenant_id": TENANT_ID, "listing_id": listing_id},
        )
        assert trends.status_code == 200
        t = trends.json()
        assert len(t["history"]) >= 80  # 接近 90 天
        assert len(t["forecast"]) >= 30
        # 历史的最新一天应早于 / 等于预测的第一天
        if t["history"] and t["forecast"]:
            assert t["history"][-1]["date"] <= t["forecast"][0]["date"]

    async def test_us_2_3_ai_explains_priority_and_suggest_for_sku(
        self, client: AsyncClient
    ) -> None:
        """US-2.3: AI 抽屉自动解释这个 SKU 为什么是这个风险等级.

        As a  运营负责人
        I want 打开 AI 抽屉就看到当前 SKU 的风险解读 + 建议采购数量
        So that 我不用自己拼凑数字,系统直接告诉我"为什么 P1"
        """
        first = await client.post(
            "/api/supplyai/skus/list",
            json={"tenant_id": TENANT_ID, "page_size": 1, "priorities": ["p1"]},
        )
        listing_id = first.json()["rows"][0]["id"]

        explain = await client.post(
            "/api/supplyai/ai/explain",
            json={"tenant_id": TENANT_ID, "listing_id": listing_id},
        )
        assert explain.status_code == 200
        body = explain.json()
        assert body["explanation"]
        assert body["status"] in {"ok", "partial", "degraded"}
        # context 必须含决策关键字段
        ctx = body["context"]
        for key in ("msku", "priority", "suggest_qty", "fba_sellable_days"):
            assert key in ctx


# ════════════════════════════════════════════════════════
# Epic 3 — 批量采购流转(草稿状态机)
# 角色:运营负责人 → 采购员
# ════════════════════════════════════════════════════════


class TestEpic3PurchaseDraftFlow:
    async def test_us_3_1_select_skus_then_generate_draft_preview(
        self, client: AsyncClient
    ) -> None:
        """US-3.1: 列表勾几个 SKU → 生成采购草稿 → DB 看到草稿.

        As a  运营负责人
        I want 在列表勾选若干建议采购的 SKU,一键生成草稿
        So that 不用一行一行手填
        """
        listed = await client.post(
            "/api/supplyai/skus/list",
            json={
                "tenant_id": TENANT_ID,
                "suggest_only": True,
                "page_size": 3,
            },
        )
        rows = listed.json()["rows"]
        assert len(rows) == 3

        items = [
            {
                "mall_id": r["mall_id"],
                "msku": r["msku"],
                "sku": r["sku"],
                "suggest_qty": r["suggest_qty"],
            }
            for r in rows
        ]
        create = await client.post(
            "/api/supplyai/purchase/draft/create",
            json={
                "tenant_id": TENANT_ID,
                "calc_run_id": rows[0]["calc_run_id"],
                "items": items,
            },
        )
        assert create.json()["created_count"] == 3

        # 立即在草稿列表能查到
        listing = await client.post(
            "/api/supplyai/purchase/draft/list",
            json={"tenant_id": TENANT_ID, "statuses": ["draft"], "page_size": 100},
        )
        assert listing.json()["total"] >= 3

    async def test_us_3_2_draft_state_machine_confirm_then_redirect_blocked(
        self, client: AsyncClient
    ) -> None:
        """US-3.2: 状态机 — confirm 后不能再 confirm,但可以 redirect.

        As a  采购员
        I want 已确认的草稿不能被重复确认,但可以转人工
        So that 流程不会越走越乱,confirmed 是可见的"承诺"
        """
        # 先建一条
        listed = await client.post(
            "/api/supplyai/skus/list",
            json={"tenant_id": TENANT_ID, "suggest_only": True, "page_size": 1},
        )
        r = listed.json()["rows"][0]
        create = await client.post(
            "/api/supplyai/purchase/draft/create",
            json={
                "tenant_id": TENANT_ID,
                "items": [
                    {
                        "mall_id": r["mall_id"],
                        "msku": r["msku"],
                        "sku": r["sku"],
                        "suggest_qty": r["suggest_qty"],
                    }
                ],
            },
        )
        draft_id = create.json()["draft_ids"][0]

        # confirm 一次成功
        first = await client.post(
            "/api/supplyai/purchase/draft/confirm",
            json={"tenant_id": TENANT_ID, "draft_id": draft_id},
        )
        assert first.status_code == 200
        assert first.json()["status"] == "confirmed"

        # confirm 第二次拒绝
        second = await client.post(
            "/api/supplyai/purchase/draft/confirm",
            json={"tenant_id": TENANT_ID, "draft_id": draft_id},
        )
        assert second.status_code == 400
        assert second.json()["detail"]["code"] == "DRAFT_INVALID_TRANSITION"

        # 但 confirmed → redirected 允许
        red = await client.post(
            "/api/supplyai/purchase/draft/redirect",
            json={"tenant_id": TENANT_ID, "draft_id": draft_id},
        )
        assert red.status_code == 200
        assert red.json()["status"] == "redirected"


# ════════════════════════════════════════════════════════
# Epic 4 — 规则配置 + 重算闭环
# 角色:规则管理员
# ════════════════════════════════════════════════════════


class TestEpic4RulesAndRecalculation:
    async def test_us_4_1_save_store_rule_then_appears_in_list(
        self, client: AsyncClient
    ) -> None:
        """US-4.1: 创建店铺级特配规则,在规则列表里能查到.

        As a  规则管理员
        I want 给店铺 1001 配一条特殊安全天数规则
        So that 该店铺的 SKU 用更激进/保守的备货策略
        """
        upsert = await client.post(
            "/api/supplyai/rules/upsert",
            json={
                "tenant_id": TENANT_ID,
                "scope_type": "store",
                "mall_id": 1001,
                "safety_days": 21,
                "purchase_duration_days": 5,
                "delivery_days": 25,
                "qc_days": 3,
                "enabled": True,
                "updated_by": "e2e-test",
            },
        )
        rule_id = upsert.json()["rule_id"]

        listed = await client.post(
            "/api/supplyai/rules/list",
            json={"tenant_id": TENANT_ID, "scope_types": ["store"]},
        )
        ids = [r["rule_id"] for r in listed.json()["rows"]]
        assert rule_id in ids

    async def test_us_4_2_recalc_after_rule_change_changes_calc_run_id(
        self, client: AsyncClient
    ) -> None:
        """US-4.2: 规则改完后跑一次 calc.run,新批次成为最新.

        As a  规则管理员
        I want 改完规则触发重算,Dashboard 立刻看到基于新规则的快照
        So that 修改不需要等一整夜的定时任务才生效
        """
        # 触发 calc.run
        run = await client.post(
            "/api/supplyai/calc/run",
            json={"tenant_id": TENANT_ID, "run_type": "rule_changed"},
        )
        new_id = run.json()["calc_run_id"]
        assert new_id.startswith("RUN-")
        assert run.json()["status"] == "success"

        # Dashboard 默认取最新 → 应该是新 ID
        snap = await client.post(
            "/api/supplyai/dashboard/snapshot", json={"tenant_id": TENANT_ID}
        )
        assert snap.json()["calc_run_id"] == new_id


# ════════════════════════════════════════════════════════
# Epic 5 — AI 助手(Agent + Tools + Foundation Skills)
# ════════════════════════════════════════════════════════


class TestEpic5AiAssistant:
    async def test_us_5_1_global_chat_returns_assistant_message(
        self, client: AsyncClient
    ) -> None:
        """US-5.1: 在 AI 抽屉问全局问题,得到 assistant 回答.

        As a  运营负责人
        I want 在 AI 抽屉自由提问"今天该关注什么"
        So that 不用自己在 Dashboard 各处拼数据
        """
        chat = await client.post(
            "/api/supplyai/ai/chat",
            json={
                "tenant_id": TENANT_ID,
                "messages": [{"role": "user", "content": "今天有哪些 SKU 风险?"}],
            },
        )
        assert chat.status_code == 200
        body = chat.json()
        assert body["role"] == "assistant"
        assert body["content"]
        assert body["status"] in {"ok", "partial", "degraded"}

    async def test_us_5_2_chat_rejects_empty_messages(
        self, client: AsyncClient
    ) -> None:
        """US-5.2: 空消息拒绝.

        防御性,避免前端 bug 把空 payload 发上来。
        """
        empty = await client.post(
            "/api/supplyai/ai/chat",
            json={"tenant_id": TENANT_ID, "messages": []},
        )
        assert empty.status_code == 400
        assert empty.json()["detail"]["code"] == "AI_EMPTY_MESSAGES"

    async def test_us_5_3_explain_unknown_sku_returns_404(
        self, client: AsyncClient
    ) -> None:
        """US-5.3: 解释不存在的 SKU,返回 404 而不是模型胡编.

        As a  调用方
        I want 工具/解释端点对不存在的资源返回明确 404
        So that 我知道是参数错而不是模型乱答
        """
        explain = await client.post(
            "/api/supplyai/ai/explain",
            json={"tenant_id": TENANT_ID, "listing_id": 99999999},
        )
        assert explain.status_code == 404
        assert explain.json()["detail"]["code"] == "SKU_NOT_FOUND"


# ════════════════════════════════════════════════════════
# Epic 6 — 数据一致性(跨端点 / 跨表)
# 角色:任何人 — 这是产品的"信任根基"
# ════════════════════════════════════════════════════════


class TestEpic6DataConsistency:
    async def test_us_6_1_same_calc_run_id_used_everywhere(
        self, client: AsyncClient
    ) -> None:
        """US-6.1: dashboard / list / detail 默认都用同一个最新 calc_run_id.

        As a  系统使用者
        I want 同一时刻打开任何页面,看到的数据都属于同一计算批次
        So that 不会出现"工作台说 12 个 P1,但列表筛 P1 只有 8 个"的错觉
        """
        snap = await client.post(
            "/api/supplyai/dashboard/snapshot", json={"tenant_id": TENANT_ID}
        )
        calc_id = snap.json()["calc_run_id"]

        listed = await client.post(
            "/api/supplyai/skus/list",
            json={"tenant_id": TENANT_ID, "page_size": 1},
        )
        assert listed.json()["rows"][0]["calc_run_id"] == calc_id

        listing_id = listed.json()["rows"][0]["id"]
        detail = await client.post(
            "/api/supplyai/skus/detail",
            json={"tenant_id": TENANT_ID, "listing_id": listing_id},
        )
        assert detail.json()["calc_run_id"] == calc_id
        assert detail.json()["summary"]["calc_run_id"] == calc_id

    async def test_us_6_2_suggest_qty_satisfies_formula_for_all_p1(
        self, client: AsyncClient
    ) -> None:
        """US-6.2: 所有 P1 SKU 的 suggest_qty 满足公式.

        suggest_qty == max(0, ceil(coverage_demand - total_stock))
        这是产品最核心的可信承诺,违反 = 系统数学不自洽。
        """
        listed = await client.post(
            "/api/supplyai/skus/list",
            json={
                "tenant_id": TENANT_ID,
                "priorities": ["p1"],
                "page_size": 50,
            },
        )
        for r in listed.json()["rows"]:
            cov = r.get("coverage_demand")
            stock = r.get("total_stock") or 0
            if cov is None:
                continue
            expected = max(0, math.ceil(cov - stock))
            assert r["suggest_qty"] == expected, (
                f"违反公式:msku={r['msku']} qty={r['suggest_qty']} "
                f"expected={expected} (cov={cov}, stock={stock})"
            )

    async def test_us_6_3_risk_level_aligned_with_fba_sellable_days(
        self, client: AsyncClient
    ) -> None:
        """US-6.3: 风险等级与 fba_sellable_days 阈值一致.

        p1 ≤ 7 / p2 8-15 / p3 16-30 / safe > 30
        """
        listed = await client.post(
            "/api/supplyai/skus/list",
            json={"tenant_id": TENANT_ID, "page_size": 48},
        )
        for r in listed.json()["rows"]:
            d = r.get("fba_sellable_days")
            if d is None:
                assert r["priority"] == "safe"
            elif d <= 7:
                assert r["priority"] == "p1", f"{r['msku']} d={d}"
            elif d <= 15:
                assert r["priority"] == "p2"
            elif d <= 30:
                assert r["priority"] == "p3"
            else:
                assert r["priority"] == "safe"

    async def test_us_6_4_forecast_daily_average_matches_snapshot(
        self, client: AsyncClient
    ) -> None:
        """US-6.4: snapshot.future_daily ≈ avg(forecast 序列),误差 ±0.5.

        快照与逐日预测同 calc_run_id 同事务写入,平均值必须自洽。
        随机抽 3 个 SKU 验证。
        """
        listed = await client.post(
            "/api/supplyai/skus/list",
            json={"tenant_id": TENANT_ID, "page_size": 3},
        )
        for r in listed.json()["rows"]:
            if r.get("future_daily") is None:
                continue
            trends = await client.post(
                "/api/supplyai/skus/trends",
                json={"tenant_id": TENANT_ID, "listing_id": r["id"]},
            )
            forecast = trends.json()["forecast"]
            if not forecast:
                continue
            avg = sum(p["qty"] for p in forecast) / len(forecast)
            assert abs(avg - r["future_daily"]) <= 0.5, (
                f"{r['msku']}: future_daily={r['future_daily']} avg(forecast)={avg}"
            )


# ════════════════════════════════════════════════════════
# Epic 7 — 跨流程贯通(最难,但最有信心)
# 一次跑完整业务路径,验证端到端不破。
# ════════════════════════════════════════════════════════


class TestEpic7EndToEndJourney:
    async def test_us_7_1_full_journey_from_workbench_to_confirmed_draft(
        self, client: AsyncClient
    ) -> None:
        """US-7.1: 完整决策路径 — 从工作台到草稿确认的全流程.

        模拟一个运营负责人上岗到下班前的完整动作:
          1. 看 Dashboard 风险分布
          2. 进备货列表筛 P1
          3. 进任一 SKU 详情看曲线
          4. 调 AI 解释
          5. 回列表勾该 SKU
          6. 生成采购草稿(预览)
          7. 进草稿页确认
          8. 验证状态变 confirmed
          9. 整个流程同一 calc_run_id 贯穿
        """
        # ① 工作台
        snap = await client.post(
            "/api/supplyai/dashboard/snapshot", json={"tenant_id": TENANT_ID}
        )
        calc_id = snap.json()["calc_run_id"]
        assert snap.json()["risk_counts"]["p1"] >= 1

        # ② 备货列表 P1
        listed = await client.post(
            "/api/supplyai/skus/list",
            json={
                "tenant_id": TENANT_ID,
                "priorities": ["p1"],
                "suggest_only": True,
                "page_size": 1,
            },
        )
        target = listed.json()["rows"][0]
        assert target["calc_run_id"] == calc_id

        # ③ SKU 详情
        detail = await client.post(
            "/api/supplyai/skus/detail",
            json={"tenant_id": TENANT_ID, "listing_id": target["id"]},
        )
        assert detail.json()["calc_run_id"] == calc_id

        # ④ AI 解释
        explain = await client.post(
            "/api/supplyai/ai/explain",
            json={"tenant_id": TENANT_ID, "listing_id": target["id"]},
        )
        assert explain.json()["explanation"]

        # ⑤ ⑥ 生成草稿
        create = await client.post(
            "/api/supplyai/purchase/draft/create",
            json={
                "tenant_id": TENANT_ID,
                "calc_run_id": calc_id,
                "items": [
                    {
                        "mall_id": target["mall_id"],
                        "msku": target["msku"],
                        "sku": target["sku"],
                        "suggest_qty": target["suggest_qty"],
                    }
                ],
            },
        )
        draft_id = create.json()["draft_ids"][0]

        # ⑦ 草稿确认
        confirm = await client.post(
            "/api/supplyai/purchase/draft/confirm",
            json={"tenant_id": TENANT_ID, "draft_id": draft_id},
        )
        assert confirm.json()["status"] == "confirmed"

        # ⑧ ⑨ 草稿与原 calc_run_id 关联
        detail_draft = await client.post(
            "/api/supplyai/purchase/draft/detail",
            json={"tenant_id": TENANT_ID, "draft_id": draft_id},
        )
        assert detail_draft.json()["calc_run_id"] == calc_id


# ════════════════════════════════════════════════════════
# Epic 8 — Excel 导出链路(技术方案 §7.6)
# 角色:运营 / 老板 — 把工作台数据导出审阅
# ════════════════════════════════════════════════════════


class TestEpic8Exports:
    async def test_us_8_1_export_filtered_p1_then_download_xlsx(
        self, client: AsyncClient
    ) -> None:
        """US-8.1: 筛 P1 → 导出 → 状态查询 → 下载 → 校验行数/列/口径.

        As a  运营负责人
        I want 把当前 P1 列表导出 xlsx 发给老板审阅
        So that 不必让老板登系统,Excel 行数/字段/筛选必须与界面一致
        """
        # 1. 列表筛 P1 — 先记录 12 行
        listed = await client.post(
            "/api/supplyai/skus/list",
            json={"tenant_id": TENANT_ID, "priorities": ["p1"], "page_size": 50},
        )
        assert listed.json()["total"] == 12
        ui_mskus = sorted(r["msku"] for r in listed.json()["rows"])

        # 2. 触发导出 — 同筛选条件
        export = await client.post(
            "/api/supplyai/exports/sku-list",
            json={"tenant_id": TENANT_ID, "priorities": ["p1"]},
        )
        assert export.status_code == 200
        task = export.json()
        assert task["row_count"] == 12  # 与列表口径一致
        assert task["status"] == "success"  # 12 行 < 5000 同步完成
        task_id = task["task_id"]

        # 3. 状态查询
        status = await client.post(
            "/api/supplyai/exports/status",
            json={"tenant_id": TENANT_ID, "task_id": task_id},
        )
        assert status.json()["status"] == "success"
        assert status.json()["row_count"] == 12

        # 4. 下载 → 解析 xlsx
        dl = await client.post(
            "/api/supplyai/exports/download",
            json={"tenant_id": TENANT_ID, "task_id": task_id},
        )
        assert dl.status_code == 200
        assert dl.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument"
        )

        wb = openpyxl.load_workbook(io.BytesIO(dl.content))
        ws = wb.active
        assert ws.max_row == 13  # 表头 + 12 行
        # 表头必含关键列
        headers = [c.value for c in ws[1]]
        for col in ("MSKU", "ASIN", "店铺", "风险等级", "建议数量"):
            assert col in headers, f"导出缺列: {col}"
        # 风险列必须全为 p1(口径与界面一致)
        risk_idx = headers.index("风险等级") + 1
        msku_idx = headers.index("MSKU") + 1
        for row in range(2, 14):
            assert ws.cell(row=row, column=risk_idx).value == "p1"
        # MSKU 集合与界面一致(顺序可不同)
        xlsx_mskus = sorted(
            ws.cell(row=row, column=msku_idx).value for row in range(2, 14)
        )
        assert xlsx_mskus == ui_mskus

    async def test_us_8_2_export_status_404_for_unknown_task(
        self, client: AsyncClient
    ) -> None:
        """US-8.2: 不存在的 task_id → 状态/下载明确 404.

        防御性,避免前端用错的 task 反复轮询无响应。
        """
        bogus = "EXP-NEVER-EXISTED"
        status = await client.post(
            "/api/supplyai/exports/status",
            json={"tenant_id": TENANT_ID, "task_id": bogus},
        )
        assert status.status_code == 404
        assert status.json()["detail"]["code"] == "EXPORT_TASK_NOT_FOUND"

        download = await client.post(
            "/api/supplyai/exports/download",
            json={"tenant_id": TENANT_ID, "task_id": bogus},
        )
        assert download.status_code == 404


# ════════════════════════════════════════════════════════
# Epic 9 — AI Agent 工具调度端到端
# (技术方案 §7.5 — 4 Tools + 二次确认)
# ════════════════════════════════════════════════════════


@dataclass
class _ScriptedClient:
    """按预设序列回放 LLM 响应,用于 e2e 模拟模型行为."""
    responses: list[Any]
    calls: list[dict] = field(default_factory=list)

    async def chat(self, messages, tools=None, max_tokens=1024, temperature=0.2):
        self.calls.append({
            "n_messages": len(messages),
            "tool_count": len(tools) if tools else 0,
        })
        return self.responses.pop(0)


class TestEpic9AiAgentTools:
    async def test_us_9_1_chat_calls_query_tool_and_overrides_tenant(
        self, client: AsyncClient, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """US-9.1: AI 对话触发 query_stockout_risk,且 tenant_id 强制对齐请求方.

        As a  开发/安全审计
        I want 模型即使主动传错的 tenant_id 也会被 Orchestrator 强制覆盖
        So that 任何模型幻觉都不会越权读他租户数据
        """
        captured_calls: list[tuple] = []
        import supplyai.domain.ai.orchestrator as orch_mod
        real_exec = orch_mod.execute_tool

        async def _capture(name, args, session):
            captured_calls.append((name, dict(args)))
            return await real_exec(name, args, session)

        monkeypatch.setattr(orch_mod, "execute_tool", _capture)

        # 模型脚本:第一轮调工具(故意传 tenant_id=1),第二轮总结
        scripted = _ScriptedClient(responses=[
            ChatResponse(
                content="",
                finish_reason="tool_calls",
                tool_calls=[ToolCall(
                    id="c1", name="query_stockout_risk",
                    arguments={"limit": 3, "tenant_id": 1},  # 错的 tenant
                )],
            ),
            ChatResponse(content="共 3 个 P1 风险 SKU。", finish_reason="stop"),
        ])
        import supplyai.api.v1.ai as ai_api
        monkeypatch.setattr(ai_api, "get_ai_client", lambda: scripted)

        chat = await client.post(
            "/api/supplyai/ai/chat",
            json={
                "tenant_id": TENANT_ID,
                "messages": [{"role": "user", "content": "哪些 SKU 紧急?"}],
            },
        )
        assert chat.status_code == 200
        assert chat.json()["finish_reason"] == "stop"
        # 工具调用了一次,且 tenant_id 被强制纠正
        assert len(captured_calls) == 1
        name, args = captured_calls[0]
        assert name == "query_stockout_risk"
        assert args["tenant_id"] == TENANT_ID, "权限边界:模型传的 tenant 必须被覆盖"

    async def test_us_9_2_ai_generate_purchase_draft_requires_confirmation(
        self, client: AsyncClient, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """US-9.2: AI 调 generate_purchase_draft 不带 confirmed=True 时不能落库.

        As a  规则审计 / 业务方
        I want AI 触发的采购动作必须经过人类二次确认才进 DB
        So that 一句"帮我下单"不会变成实际采购指令
        """
        # 取一个真实 P1 用作模型构造的 items
        listed = await client.post(
            "/api/supplyai/skus/list",
            json={"tenant_id": TENANT_ID, "priorities": ["p1"], "page_size": 1},
        )
        target = listed.json()["rows"][0]

        scripted = _ScriptedClient(responses=[
            ChatResponse(
                content="",
                finish_reason="tool_calls",
                tool_calls=[ToolCall(
                    id="c1", name="generate_purchase_draft",
                    arguments={
                        "items": [{
                            "msku": target["msku"],
                            "mall_id": target["mall_id"],
                            "suggest_qty": target["suggest_qty"],
                        }],
                        "confirmed": False,  # 关键:未确认
                    },
                )],
            ),
            ChatResponse(
                content="预览已准备,请人工确认 SKU、数量、供应商三项后再下单。",
                finish_reason="stop",
            ),
        ])
        import supplyai.api.v1.ai as ai_api
        monkeypatch.setattr(ai_api, "get_ai_client", lambda: scripted)

        # 调用前先确认 DB 没草稿
        before = await client.post(
            "/api/supplyai/purchase/draft/list",
            json={"tenant_id": TENANT_ID, "page_size": 100},
        )
        before_count = before.json()["total"]

        chat = await client.post(
            "/api/supplyai/ai/chat",
            json={
                "tenant_id": TENANT_ID,
                "messages": [{"role": "user", "content": "帮我下采购单"}],
            },
        )
        assert chat.status_code == 200

        # 关键断言:草稿数量没变 — 模型未 confirmed,工具拦截
        after = await client.post(
            "/api/supplyai/purchase/draft/list",
            json={"tenant_id": TENANT_ID, "page_size": 100},
        )
        assert after.json()["total"] == before_count, "未确认就不能落库"


# ════════════════════════════════════════════════════════
# Epic 10 — FBM 兜底 + 数据质量(技术方案 §11)
# ════════════════════════════════════════════════════════


@pytest_asyncio.fixture
async def fbm_listing():
    """临时插入一条 FBM listing,跑完测试自动清理."""
    listing_id = 999001
    async with async_session_factory() as session:
        session.add(MkListingProductSources(
            tenant_id=TENANT_ID,
            listing_id=listing_id,
            mall_id=1001,
            msku="FBM-TEST-001",
            sku="FBM-TEST-SKU-US",
            asin="B0FBMTEST",
            delivery_method="FBM",
            listing_status="active",
            product_name="FBM 测试商品 — E2E 临时数据",
            country_code="US",
        ))
        await session.commit()
    yield listing_id
    async with async_session_factory() as session:
        await session.execute(
            delete(MkListingProductSources).where(
                MkListingProductSources.listing_id == listing_id,
            )
        )
        await session.commit()


class TestEpic10FbmAndDataQuality:
    async def test_us_10_1_fbm_skus_excluded_from_list_but_detail_explains(
        self, client: AsyncClient, fbm_listing: int
    ) -> None:
        """US-10.1: FBM 商品不进备货列表,但详情有明确兜底而不是 SKU_NOT_FOUND.

        As a  混卖 FBA+FBM 的运营
        I want 列表只显示 FBA(因为只有 FBA 才需要算备货),
              但点开 FBM 详情能看到"该商品当前 Phase 不支持备货分析"
        So that 我不会以为系统漏数据,理解到这是产品 Phase 限制
        """
        # 列表绝不含 FBM(基于 stat 表,calc.run 只跑 FBA)
        listed = await client.post(
            "/api/supplyai/skus/list",
            json={"tenant_id": TENANT_ID, "page_size": 100},
        )
        for r in listed.json()["rows"]:
            assert r["delivery_method"] == "FBA"
        msku_set = {r["msku"] for r in listed.json()["rows"]}
        assert "FBM-TEST-001" not in msku_set

        # 详情走 FBM listing → 返 FBM_NOT_SUPPORTED 而不是 SKU_NOT_FOUND
        detail = await client.post(
            "/api/supplyai/skus/detail",
            json={"tenant_id": TENANT_ID, "listing_id": fbm_listing},
        )
        assert detail.status_code == 400
        assert detail.json()["detail"]["code"] == "FBM_NOT_SUPPORTED"

    async def test_us_10_2_detail_data_quality_does_not_silently_fill_zero(
        self, client: AsyncClient
    ) -> None:
        """US-10.2: detail.data_quality 必须返回结构化警告,不能用 0 静默掩盖缺失.

        As a  数据审计员
        I want 当某 SKU 的 unit_cost / forecast 缺失时,
              detail.data_quality.missing_fields 必须列出来,不是默认填 0
        So that 我知道这条数据不可信,不会被假完整骗
        """
        listed = await client.post(
            "/api/supplyai/skus/list",
            json={"tenant_id": TENANT_ID, "page_size": 1},
        )
        listing_id = listed.json()["rows"][0]["id"]

        detail = await client.post(
            "/api/supplyai/skus/detail",
            json={"tenant_id": TENANT_ID, "listing_id": listing_id},
        )
        d = detail.json()
        # 结构必含 data_quality 三件套
        assert "data_quality" in d
        dq = d["data_quality"]
        assert "missing_fields" in dq
        assert "warnings" in dq
        # 类型契约
        assert isinstance(dq["missing_fields"], list)
        assert isinstance(dq["warnings"], list)
        # 任一警告必须有 code/message,不能空字符串糊弄
        for w in dq["warnings"]:
            assert w.get("code"), "warning 缺 code"
            assert w.get("message"), "warning 缺 message"
            assert w.get("severity") in {"info", "warn", "error"}


# ════════════════════════════════════════════════════════
# Epic 11 — 规则配置真正影响计算结果
# ════════════════════════════════════════════════════════


class TestEpic11RulesActuallyAffectCalc:
    async def test_us_11_1_store_rule_changes_safety_days_after_recalc(
        self, client: AsyncClient
    ) -> None:
        """US-11.1: 改店铺规则 safety_days,calc.run 后该店铺 SKU 真的用新值.

        As a  规则管理员
        I want 调整某店铺的 safety_days 后,该店铺所有 SKU 重算后的
              建议采购量 / coverage_demand 真的反映新规则
        So that 改规则不是"前端假动作",而是端到端起作用
        """
        # 1. 取一个 mall_id=1001 的 SKU,记录当前 safety_days
        listed = await client.post(
            "/api/supplyai/skus/list",
            json={"tenant_id": TENANT_ID, "mall_ids": [1001], "page_size": 1},
        )
        rows = listed.json()["rows"]
        assert rows, "seed 应有 mall_id=1001 的 SKU"
        target_msku = rows[0]["msku"]
        before = await client.post(
            "/api/supplyai/skus/list",
            json={
                "tenant_id": TENANT_ID,
                "mall_ids": [1001],
                "keyword": target_msku,
                "page_size": 1,
            },
        )
        before_safety = before.json()["rows"][0].get("safety_days") or 14

        # 2. 创建店铺级规则 — 把 safety_days 改成与之前显著不同的值
        new_safety = 99 if before_safety != 99 else 88
        await client.post(
            "/api/supplyai/rules/upsert",
            json={
                "tenant_id": TENANT_ID,
                "scope_type": "store",
                "mall_id": 1001,
                "safety_days": new_safety,
                "purchase_duration_days": 5,
                "delivery_days": 20,
                "qc_days": 2,
                "enabled": True,
                "updated_by": "e2e",
            },
        )

        # 3. 触发重算
        run = await client.post(
            "/api/supplyai/calc/run",
            json={"tenant_id": TENANT_ID, "run_type": "rule_changed"},
        )
        new_run_id = run.json()["calc_run_id"]

        # 4. 同 SKU 在新批次的 safety_days 必须 = new_safety
        after = await client.post(
            "/api/supplyai/skus/list",
            json={
                "tenant_id": TENANT_ID,
                "calc_run_id": new_run_id,
                "mall_ids": [1001],
                "keyword": target_msku,
                "page_size": 1,
            },
        )
        after_row = after.json()["rows"][0]
        assert after_row["safety_days"] == new_safety, (
            f"规则没生效:期待 safety_days={new_safety},实得 {after_row['safety_days']}"
        )

    async def test_us_11_2_sku_rule_overrides_store_then_disable_falls_back(
        self, client: AsyncClient
    ) -> None:
        """US-11.2: SKU 级规则覆盖店铺级,disable SKU 规则后回退到店铺级.

        三层优先级:sku > store > global > default(技术方案 §4.3)
        disable 一条规则,Calc Engine 必须忽略它,回退到下一层。
        """
        # 取一个 mall_id=1001 的 SKU
        listed = await client.post(
            "/api/supplyai/skus/list",
            json={"tenant_id": TENANT_ID, "mall_ids": [1001], "page_size": 1},
        )
        target = listed.json()["rows"][0]
        target_msku = target["msku"]

        # ① 创建 store 规则 safety_days=20
        await client.post(
            "/api/supplyai/rules/upsert",
            json={
                "tenant_id": TENANT_ID,
                "scope_type": "store",
                "mall_id": 1001,
                "safety_days": 20,
                "purchase_duration_days": 5,
                "delivery_days": 20,
                "qc_days": 2,
                "enabled": True,
            },
        )
        # ② 创建 sku 规则 safety_days=5(更具体,优先级最高)
        sku_rule = await client.post(
            "/api/supplyai/rules/upsert",
            json={
                "tenant_id": TENANT_ID,
                "scope_type": "sku",
                "mall_id": 1001,
                "msku": target_msku,
                "safety_days": 5,
                "purchase_duration_days": 5,
                "delivery_days": 20,
                "qc_days": 2,
                "enabled": True,
            },
        )
        sku_rule_id = sku_rule.json()["rule_id"]

        # ③ calc.run → 该 SKU safety_days 应当 = 5(SKU 级覆盖)
        run1 = await client.post(
            "/api/supplyai/calc/run",
            json={"tenant_id": TENANT_ID, "run_type": "rule_changed"},
        )
        list1 = await client.post(
            "/api/supplyai/skus/list",
            json={
                "tenant_id": TENANT_ID,
                "calc_run_id": run1.json()["calc_run_id"],
                "keyword": target_msku,
                "page_size": 1,
            },
        )
        assert list1.json()["rows"][0]["safety_days"] == 5, "sku 级规则未覆盖 store"

        # ④ disable SKU 级规则 → calc.run 应回退到 store 的 20
        disable = await client.post(
            "/api/supplyai/rules/disable",
            json={"tenant_id": TENANT_ID, "rule_id": sku_rule_id},
        )
        assert disable.json()["enabled"] is False

        run2 = await client.post(
            "/api/supplyai/calc/run",
            json={"tenant_id": TENANT_ID, "run_type": "rule_changed"},
        )
        list2 = await client.post(
            "/api/supplyai/skus/list",
            json={
                "tenant_id": TENANT_ID,
                "calc_run_id": run2.json()["calc_run_id"],
                "keyword": target_msku,
                "page_size": 1,
            },
        )
        assert list2.json()["rows"][0]["safety_days"] == 20, "disable 后未回退到 store"


# ════════════════════════════════════════════════════════
# Epic 12 — API 全 POST 契约(技术方案 §6)
# ════════════════════════════════════════════════════════


class TestEpic12ApiContract:
    async def test_us_12_1_all_business_routes_are_post_only(
        self, client: AsyncClient
    ) -> None:
        """US-12.1: 业务 API 全 POST,不允许 GET / PUT / DELETE.

        As a  接口审计 / 前端 SDK 作者
        I want /api/supplyai/* 业务路由清一色 POST(除文档/health 类元路由)
        So that SDK 不必区分 method,前端不会因为 method 写错而 405
        """
        from supplyai.main import app

        # 允许 GET 的元路由白名单
        meta_get_paths = {"/", "/api/supplyai/_health"}
        # FastAPI 自动注入的(路径不在 /api/supplyai/* 下,不参与契约)
        excluded_prefixes = ("/docs", "/redoc", "/openapi.json")

        violations: list[tuple[str, set[str]]] = []
        for route in app.routes:
            path = getattr(route, "path", "")
            methods = set(getattr(route, "methods", set()) or set())
            methods.discard("HEAD")  # FastAPI 给 GET 路由自动加 HEAD
            methods.discard("OPTIONS")  # CORS 预检
            if not path or path in meta_get_paths:
                continue
            if any(path.startswith(p) for p in excluded_prefixes):
                continue
            if not path.startswith("/api/supplyai/"):
                continue
            if methods != {"POST"}:
                violations.append((path, methods))

        assert not violations, (
            "以下业务路由违反全 POST 契约:\n"
            + "\n".join(f"  {p}: {m}" for p, m in violations)
        )
